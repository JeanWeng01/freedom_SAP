"""Read and validate SAP_bot.xlsx."""

import os
import logging
from dataclasses import dataclass

import openpyxl

log = logging.getLogger(__name__)

# Expected column layout (1-indexed) after rearrangement:
#   A(1): Document_1   B(2): Invoice       C(3): POD_Filename
#   D(4): Charge_Hhrs  E(5): Charges        F(6): Amount
#   G(7): POD_Base_Path
#   H-J: reserved
#   K(11): Document_2  L(12): Charge_Hhrs   M(13): Amount  N(14): Charges
EXPECTED_COLUMNS = {
    1:  "Document_1",
    2:  "Invoice",
    3:  "POD_Filename",
    4:  "Charge_Hhrs",
    5:  "Charges",
    6:  "Amount",
    7:  "POD_Base_Path",
    11: "Document_2",
    12: "Charge_Hhrs",
    13: "Amount",
    14: "Charges",
}


@dataclass
class InvoiceRow:
    """One row from SAP_bot.xlsx."""
    row_number: int
    document_1: str
    invoice: str
    pod_filename: str               # bare filename, no extension (col C)
    charge_hours_1: float | None    # col D
    charge_type_1: str | None       # col E — e.g. "Waiting Charges" or None
    charge_amount_1: float | None   # col F — computed dollar amount
    pod_base_path: str              # col G
    document_2: str | None          # col K
    charge_hours_2: float | None    # col L
    charge_amount_2: float | None   # col M
    charge_type_2: str | None       # col N

    @property
    def is_collective(self) -> bool:
        return bool(self.document_2)

    @property
    def has_charges_leg1(self) -> bool:
        return bool(self.charge_type_1)

    @property
    def has_charges_leg2(self) -> bool:
        return bool(self.charge_type_2)

    @property
    def pod_full_path(self) -> str:
        return os.path.join(self.pod_base_path, self.pod_filename + ".pdf")


def validate_headers(ws) -> list[str]:
    """Check that expected columns exist. Returns list of error messages (empty = OK)."""
    errors = []
    for col_num, expected_name in EXPECTED_COLUMNS.items():
        actual = ws.cell(row=1, column=col_num).value
        if actual is None:
            errors.append(f"Column {col_num} is empty — expected '{expected_name}'")
        elif str(actual).strip() != expected_name:
            errors.append(
                f"Column {col_num}: expected '{expected_name}', got '{actual}'"
            )
    return errors


def _cell_val(ws, row: int, col: int):
    """Get cell value, returning None for blank cells."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def read_excel(path: str) -> list[InvoiceRow]:
    """Read SAP_bot.xlsx and return a list of InvoiceRow objects.

    Raises SystemExit if file is missing or headers are invalid.
    """
    if not os.path.isfile(path):
        log.error("Excel file not found: %s", path)
        raise SystemExit(f"Excel file not found: {path}")

    # Try data_only first for cached formula values; also load raw for fallback
    wb_data = openpyxl.load_workbook(path, data_only=True)
    wb_raw = openpyxl.load_workbook(path, data_only=False)
    ws = wb_data.active
    ws_raw = wb_raw.active

    # Validate headers (use raw workbook — headers are always plain text)
    header_errors = validate_headers(ws_raw)
    if header_errors:
        for err in header_errors:
            log.error("Excel header error: %s", err)
        raise SystemExit(
            "Excel header validation failed:\n  " + "\n  ".join(header_errors)
        )

    rows = []
    for row_num in range(2, ws.max_row + 1):
        doc1 = _cell_val(ws, row_num, 1)
        if doc1 is None:
            continue  # skip blank rows

        charge_hrs_1 = _cell_val(ws, row_num, 4)    # col D
        charge_hrs_2 = _cell_val(ws, row_num, 12)   # col L

        # Amount: prefer cached formula value; if None, compute from hours
        HOURLY_RATE = 42.84
        charge_amt_1 = _cell_val(ws, row_num, 6)    # col F cached
        if charge_amt_1 is None and charge_hrs_1 is not None:
            charge_amt_1 = round(float(charge_hrs_1) * HOURLY_RATE, 2)
            log.debug("Computed charge_amount_1 for row %d: %s", row_num, charge_amt_1)

        charge_amt_2 = _cell_val(ws, row_num, 13)   # col M cached
        if charge_amt_2 is None and charge_hrs_2 is not None:
            charge_amt_2 = round(float(charge_hrs_2) * HOURLY_RATE, 2)
            log.debug("Computed charge_amount_2 for row %d: %s", row_num, charge_amt_2)

        inv = InvoiceRow(
            row_number=row_num,
            document_1=str(doc1).strip(),
            invoice=str(_cell_val(ws, row_num, 2) or "").strip(),
            pod_filename=str(_cell_val(ws, row_num, 3) or "").strip(),
            charge_hours_1=float(charge_hrs_1) if charge_hrs_1 is not None else None,
            charge_type_1=_cell_val(ws, row_num, 5),    # col E
            charge_amount_1=float(charge_amt_1) if charge_amt_1 is not None else None,
            pod_base_path=str(_cell_val(ws, row_num, 7) or "").strip(),
            document_2=str(_cell_val(ws, row_num, 11)).strip() if _cell_val(ws, row_num, 11) else None,
            charge_hours_2=float(charge_hrs_2) if charge_hrs_2 is not None else None,
            charge_amount_2=float(charge_amt_2) if charge_amt_2 is not None else None,
            charge_type_2=_cell_val(ws, row_num, 14),   # col N
        )
        rows.append(inv)

    log.info("Loaded %d rows from %s", len(rows), path)
    wb_data.close()
    wb_raw.close()
    return rows
